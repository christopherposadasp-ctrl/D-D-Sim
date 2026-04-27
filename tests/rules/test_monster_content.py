from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.content.enemies import (
    BENCHMARK_ENEMY_PRESET_ID_BY_VARIANT,
    BENCHMARK_MONSTER_COUNTS,
    MONSTER_DEFINITIONS,
    create_enemy,
    get_enemy_preset,
    get_unit_bonus_action_ids,
    get_unit_reaction_ids,
    unit_has_creature_tag,
    unit_is_undead,
)
from backend.content.special_actions import DRAGON_BREATH_ACTIONS, LEGENDARY_CONE_FEAR_ACTIONS, LEGENDARY_SPHERE_ACTIONS
from backend.engine.models.state import WeaponDamageComponent, WeaponProfile
from tests.rules.monster_expectations import MONSTER_EXPECTATIONS, REMAINING_MONSTER_IDS


REPO_ROOT = Path(__file__).resolve().parents[2]
SRD_CREATURES_WORKBOOK = REPO_ROOT / "docs" / "reference" / "srd_Creatures_Trimmed_V4.xlsx"


def serialize_damage_components(
    components: list[WeaponDamageComponent] | tuple[WeaponDamageComponent, ...] | None,
) -> tuple[tuple[str, tuple[tuple[int, int], ...], int], ...]:
    if not components:
        return ()
    return tuple(
        (
            component.damage_type,
            tuple((die.count, die.sides) for die in component.damage_dice),
            component.damage_modifier,
        )
        for component in components
    )


def serialize_weapon_profile(weapon: WeaponProfile) -> dict[str, object]:
    return {
        "kind": weapon.kind,
        "attack_bonus": weapon.attack_bonus,
        "ability_modifier": weapon.ability_modifier,
        "damage_modifier": weapon.damage_modifier,
        "damage_type": weapon.damage_type,
        "selectable_damage_types": tuple(weapon.selectable_damage_types or ()),
        "reach": weapon.reach,
        "range": None if weapon.range is None else (weapon.range.normal, weapon.range.long),
        "damage_dice": tuple((die.count, die.sides) for die in weapon.damage_dice),
        "advantage_damage_dice": tuple((die.count, die.sides) for die in weapon.advantage_damage_dice or ()),
        "damage_components": serialize_damage_components(weapon.damage_components),
        "advantage_damage_components": serialize_damage_components(weapon.advantage_damage_components),
        "advantage_against_self_grappled_target": bool(weapon.advantage_against_self_grappled_target),
    }


@pytest.mark.parametrize("variant_id", REMAINING_MONSTER_IDS)
def test_remaining_monster_roster_matches_expectation_table(variant_id: str) -> None:
    expectation = MONSTER_EXPECTATIONS[variant_id]
    definition = MONSTER_DEFINITIONS[variant_id]
    preset = get_enemy_preset(expectation.benchmark_preset_id)

    assert BENCHMARK_ENEMY_PRESET_ID_BY_VARIANT[variant_id] == expectation.benchmark_preset_id
    assert BENCHMARK_MONSTER_COUNTS[variant_id] == expectation.benchmark_count
    assert BENCHMARK_MONSTER_COUNTS[variant_id] == max(1, min(15, round(100 / definition.max_hp)))

    assert preset.preset_id == expectation.benchmark_preset_id
    assert len(preset.units) == expectation.benchmark_count
    assert all(unit.variant_id == variant_id for unit in preset.units)

    assert definition.ai_profile_id == expectation.ai_profile_id
    assert definition.max_hp == expectation.max_hp
    assert definition.ac == expectation.ac
    assert definition.speed == expectation.speed
    assert definition.initiative_mod == expectation.initiative_mod
    assert definition.passive_perception == expectation.passive_perception
    assert definition.size_category == expectation.size_category
    assert (definition.footprint.width, definition.footprint.height) == expectation.footprint
    assert definition.ability_mods.model_dump() == expectation.ability_mods
    assert tuple(definition.trait_ids) == expectation.trait_ids
    assert tuple(definition.role_tags) == expectation.role_tags
    assert tuple(definition.special_action_ids) == expectation.special_action_ids
    assert tuple(definition.dragon_breath_profile_ids.items()) == expectation.dragon_breath_profile_ids
    assert tuple(definition.legendary_action_ids) == expectation.legendary_action_ids
    assert tuple(definition.damage_resistances) == expectation.damage_resistances
    assert tuple(definition.damage_immunities) == expectation.damage_immunities
    assert tuple(definition.damage_vulnerabilities) == expectation.damage_vulnerabilities
    assert tuple(definition.condition_immunities) == expectation.condition_immunities
    assert tuple(definition.creature_tags) == expectation.creature_tags
    assert tuple(definition.movement_modes) == expectation.movement_modes
    runtime_unit = create_enemy("E1", variant_id)
    assert tuple(get_unit_bonus_action_ids(runtime_unit)) == expectation.bonus_action_ids
    assert tuple(get_unit_reaction_ids(runtime_unit)) == expectation.reaction_ids
    assert tuple(runtime_unit.creature_tags) == expectation.creature_tags
    assert tuple(runtime_unit.condition_immunities) == expectation.condition_immunities
    assert tuple(runtime_unit.movement_modes) == expectation.movement_modes
    assert tuple(runtime_unit.role_tags) == expectation.role_tags
    for pool_id, expected_uses in expectation.resource_pools:
        assert runtime_unit.resource_pools.get(pool_id) == expected_uses
    assert unit_has_creature_tag(runtime_unit, "undead") is ("undead" in expectation.creature_tags)
    assert unit_is_undead(runtime_unit) is ("undead" in expectation.creature_tags)

    assert set(definition.attacks) == set(expectation.attacks)
    for weapon_id, attack_expectation in expectation.attacks.items():
        assert serialize_weapon_profile(definition.attacks[weapon_id]) == {
            "kind": attack_expectation.kind,
            "attack_bonus": attack_expectation.attack_bonus,
            "ability_modifier": attack_expectation.ability_modifier,
            "damage_modifier": attack_expectation.damage_modifier,
            "damage_type": attack_expectation.damage_type,
            "selectable_damage_types": attack_expectation.selectable_damage_types,
            "reach": attack_expectation.reach,
            "range": attack_expectation.range,
            "damage_dice": attack_expectation.damage_dice,
            "advantage_damage_dice": attack_expectation.advantage_damage_dice,
            "damage_components": tuple(
                (component.damage_type, component.damage_dice, component.damage_modifier)
                for component in attack_expectation.damage_components
            ),
            "advantage_damage_components": tuple(
                (component.damage_type, component.damage_dice, component.damage_modifier)
                for component in attack_expectation.advantage_damage_components
            ),
            "advantage_against_self_grappled_target": attack_expectation.advantage_against_self_grappled_target,
        }

    if "pack_tactics" in expectation.special_mechanics:
        assert "pack_tactics" in definition.trait_ids

    if "aura_of_authority" in expectation.special_mechanics:
        assert "aura_of_authority" in definition.trait_ids

    if "sunlight_sensitivity" in expectation.special_mechanics:
        assert "sunlight_sensitivity" in definition.trait_ids

    if "selectable_damage" in expectation.special_mechanics:
        chromatic_bolt = definition.attacks["chromatic_bolt"]
        assert tuple(chromatic_bolt.selectable_damage_types or ()) == (
            "acid",
            "cold",
            "fire",
            "lightning",
            "poison",
            "thunder",
        )

    if "advantage_poison" in expectation.special_mechanics:
        longbow = definition.attacks["longbow"]
        assert serialize_damage_components(longbow.advantage_damage_components) == (("poison", ((3, 4),), 0),)

    if "advantage_damage" in expectation.special_mechanics:
        for weapon_id in ("scimitar", "shortbow"):
            weapon = definition.attacks[weapon_id]
            assert tuple((die.count, die.sides) for die in weapon.advantage_damage_dice or ()) == ((1, 4),)

    if "split_damage" in expectation.special_mechanics:
        ritual_sickle = definition.attacks["ritual_sickle"]
        assert serialize_damage_components(ritual_sickle.damage_components) == (
            ("slashing", ((1, 4),), 1),
            ("necrotic", (), 1),
        )

    if "harry_target" in expectation.special_mechanics:
        bite = definition.attacks["bite"]
        assert bite.on_hit_effects is not None
        assert [effect.kind for effect in bite.on_hit_effects] == ["harry_target"]

    if "prone_on_hit" in expectation.special_mechanics:
        weapon_id = "bite"
        if variant_id == "brown_bear":
            weapon_id = "claw"
        elif variant_id == "tiger":
            weapon_id = "rend"
        elif variant_id == "ankylosaurus":
            weapon_id = "tail"
        elif variant_id == "warhorse_skeleton":
            weapon_id = "hooves"
        weapon = definition.attacks[weapon_id]
        assert weapon.on_hit_effects is not None
        assert [effect.kind for effect in weapon.on_hit_effects] == ["prone_on_hit"]

    if "multiattack" in expectation.special_mechanics:
        multiattack = next(action for action in definition.attack_actions if action.action_id == "multiattack")
        assert definition.default_melee_attack_action_id == "multiattack"
        expected_step_count = 3 if "triple_multiattack" in expectation.special_mechanics else 2
        assert len(multiattack.steps) == expected_step_count

    if "parry" in expectation.special_mechanics:
        assert tuple(get_unit_reaction_ids(runtime_unit)) == ("opportunity_attack", "parry")

    if "redirect_attack" in expectation.special_mechanics:
        assert tuple(get_unit_reaction_ids(runtime_unit)) == ("opportunity_attack", "redirect_attack")

    if "bloodied_frenzy" in expectation.special_mechanics:
        assert "bloodied_frenzy" in definition.trait_ids

    if "rampage" in expectation.special_mechanics:
        assert tuple(get_unit_bonus_action_ids(runtime_unit)) == ("rampage",)
        assert runtime_unit.resource_pools.get("rampage_uses") == 1

    if "nimble_escape" in expectation.special_mechanics:
        assert tuple(get_unit_bonus_action_ids(runtime_unit)) == ("disengage",)

    if "undead_fortitude" in expectation.special_mechanics:
        assert "undead_fortitude" in definition.trait_ids

    if "ice_walk" in expectation.special_mechanics:
        assert "ice_walk" in definition.trait_ids

    if "legendary_resistance" in expectation.special_mechanics:
        assert "legendary_resistance" in definition.trait_ids
        assert runtime_unit.resource_pools.get("legendary_resistance_uses") == 3

    if "pounce" in expectation.special_mechanics:
        assert "pounce" in definition.legendary_action_ids
        assert runtime_unit.resource_pools.get("legendary_action_uses") == 3

    if "freezing_burst" in expectation.special_mechanics:
        burst = LEGENDARY_SPHERE_ACTIONS["freezing_burst"]
        assert "freezing_burst" in definition.legendary_action_ids
        assert runtime_unit.resource_pools.get("freezing_burst_available") == 1
        assert burst.resource_pool_id == "freezing_burst_available"
        assert burst.save_ability == "con"
        assert burst.save_dc == 14
        assert burst.range_squares == 24
        assert burst.radius_squares == 6
        assert burst.damage_die_count == 2
        assert burst.damage_die_sides == 6
        assert burst.damage_type == "cold"

    if "frightful_presence" in expectation.special_mechanics:
        fear = LEGENDARY_CONE_FEAR_ACTIONS["frightful_presence"]
        assert "frightful_presence" in definition.legendary_action_ids
        assert runtime_unit.resource_pools.get("frightful_presence_available") == 1
        assert fear.resource_pool_id == "frightful_presence_available"
        assert fear.save_ability == "wis"
        assert fear.save_dc == 14
        assert fear.range_squares == 6
        assert fear.duration_rounds == 10
        assert burst.speed_zero_on_failed_save is True

    if "cold_breath" in expectation.special_mechanics:
        assert "cold_breath" in definition.special_action_ids
        assert runtime_unit.resource_pools.get("cold_breath_available") == 1

    if "adult_cold_breath" in expectation.special_mechanics:
        breath = DRAGON_BREATH_ACTIONS["adult_white_cold_breath"]
        assert breath.action_id == "cold_breath"
        assert breath.resource_pool_id == "cold_breath_available"
        assert breath.save_ability == "con"
        assert breath.save_dc == 19
        assert breath.range_squares == 12
        assert breath.damage_die_count == 12
        assert breath.damage_die_sides == 8
        assert breath.damage_type == "cold"
        assert breath.recharge_threshold == 5

    if "adult_fire_breath" in expectation.special_mechanics:
        breath = DRAGON_BREATH_ACTIONS["adult_red_fire_breath"]
        assert breath.action_id == "fire_breath"
        assert breath.resource_pool_id == "fire_breath_available"
        assert breath.save_ability == "dex"
        assert breath.save_dc == 21
        assert breath.range_squares == 12
        assert breath.damage_die_count == 17
        assert breath.damage_die_sides == 6
        assert breath.damage_type == "fire"
        assert breath.recharge_threshold == 5

    if "fire_breath" in expectation.special_mechanics:
        assert "fire_breath" in definition.special_action_ids
        assert runtime_unit.resource_pools.get("fire_breath_available") == 1

    if "no_legendary_actions" in expectation.special_mechanics:
        legendary_fields = (
            definition.action_ids,
            definition.special_action_ids,
            definition.bonus_action_ids,
            definition.reaction_ids,
            definition.trait_ids,
            tuple(pool_id for pool_id, _uses in expectation.resource_pools),
        )
        assert not any("legendary" in item for field in legendary_fields for item in field)

    if "grappled_target_advantage" in expectation.special_mechanics:
        hammer = definition.attacks["light_hammer"]
        thrown_hammer = definition.attacks["light_hammer_throw"]
        assert hammer.advantage_against_self_grappled_target is True
        assert thrown_hammer.advantage_against_self_grappled_target is True

    if "grapple_on_hit" in expectation.special_mechanics:
        grapple_expectations = {
            "bugbear_warrior": ("grab", 12),
            "giant_crab": ("claw", 11),
            "grick": ("tentacles", 12),
            "griffon": ("rend", 14),
        }
        weapon_id, expected_escape_dc = grapple_expectations[variant_id]
        weapon = definition.attacks[weapon_id]
        assert weapon.on_hit_effects is not None
        assert [(effect.kind, effect.max_target_size, effect.escape_dc) for effect in weapon.on_hit_effects] == [
            ("grapple_on_hit", "medium", expected_escape_dc)
        ]

    if "opening_flight_landing" in expectation.special_mechanics:
        assert "opening_flight_landing" in definition.trait_ids
        assert runtime_unit.resource_pools.get("opening_landing_uses") == 1

    if "prone_on_hit" in expectation.special_mechanics and variant_id in {
        "brown_bear",
        "tiger",
        "mastiff",
        "ankylosaurus",
        "warhorse_skeleton",
    }:
        if variant_id == "brown_bear":
            weapon_id = "claw"
        elif variant_id == "tiger":
            weapon_id = "rend"
        elif variant_id == "mastiff":
            weapon_id = "bite"
        elif variant_id == "warhorse_skeleton":
            weapon_id = "hooves"
        else:
            weapon_id = "tail"
        expected_size = (
            "huge"
            if variant_id == "ankylosaurus"
            else "large"
            if variant_id in {"brown_bear", "tiger", "warhorse_skeleton"}
            else "medium"
        )
        weapon = definition.attacks[weapon_id]
        assert weapon.on_hit_effects is not None
        assert [(effect.kind, effect.max_target_size) for effect in weapon.on_hit_effects] == [
            ("prone_on_hit", expected_size)
        ]


@pytest.mark.parametrize("creature_name", ("Young White Dragon", "Young Red Dragon", "Adult White Dragon"))
def test_completed_dragon_workbook_rows_are_marked_modeled(creature_name: str) -> None:
    workbook = load_workbook(SRD_CREATURES_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    creature_index = headers.index("Creature")
    modeled_index = headers.index("Modeled")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[creature_index] == creature_name:
            assert row[modeled_index] == "Yes"
            return

    raise AssertionError(f"{creature_name} row not found in SRD creature workbook")


def test_incomplete_adult_red_dragon_workbook_row_remains_unmodeled() -> None:
    workbook = load_workbook(SRD_CREATURES_WORKBOOK, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    creature_index = headers.index("Creature")
    modeled_index = headers.index("Modeled")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[creature_index] == "Adult Red Dragon":
            assert row[modeled_index] != "Yes"
            return

    raise AssertionError("Adult Red Dragon row not found in SRD creature workbook")
